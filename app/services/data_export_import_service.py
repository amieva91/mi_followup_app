"""
Servicio de exportación e importación de datos de usuario en formato Excel (.xlsx).
Permite backup y restauración completa de datos.
"""
from io import BytesIO
from datetime import datetime, date
from typing import Dict, Any, List, Tuple, Optional

from app import db
from app.models import (
    User, Broker, BrokerAccount, Asset, Transaction, CashFlow, PortfolioHolding,
    IncomeCategory, Income, ExpenseCategory, Expense, DebtPlan,
    Bank, BankBalance, RealEstateProperty, PropertyValuation,
    Watchlist, WatchlistConfig,
)


def export_user_data(user_id: int) -> BytesIO:
    """Exporta todos los datos del usuario a un archivo .xlsx en memoria."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # Quitar hoja por defecto

    # --- Hoja Transacciones (acciones, ETFs, crypto, metales) ---
    ws_txn = wb.create_sheet("Transacciones", 0)
    txns = Transaction.query.filter_by(user_id=user_id).order_by(Transaction.transaction_date).all()
    ws_txn.append([
        "broker", "cuenta", "tipo_activo", "symbol", "isin", "tipo", "fecha", "cantidad",
        "precio", "importe", "divisa", "comision", "fees", "tax", "descripcion", "notas", "origen"
    ])
    for t in txns:
        asset = t.asset
        acct = t.account
        broker = acct.broker.name if acct and acct.broker else ""
        account_name = acct.account_name if acct else ""
        asset_type = asset.asset_type if asset else ""
        symbol = asset.symbol or "" if asset else ""
        isin = asset.isin or "" if asset else ""
        fecha = t.transaction_date.strftime("%Y-%m-%d") if t.transaction_date else ""
        ws_txn.append([
            broker, account_name, asset_type, symbol, isin, t.transaction_type,
            fecha, t.quantity, t.price, t.amount, t.currency or "EUR",
            t.commission or 0, t.fees or 0, t.tax or 0,
            (t.description or "")[:500], (t.notes or "")[:500], t.source or "MANUAL"
        ])

    cf_rows = CashFlow.query.filter_by(user_id=user_id).order_by(CashFlow.flow_date).all()
    ws_txn.append([])
    ws_txn.append(["--- CashFlow (depósitos/retiradas) ---"])
    ws_txn.append(["broker", "cuenta", "tipo", "fecha", "importe", "divisa", "descripcion"])
    for cf in cf_rows:
        acct = cf.account
        broker = acct.broker.name if acct and acct.broker else ""
        account_name = acct.account_name if acct else ""
        ws_txn.append([
            broker, account_name, cf.flow_type,
            cf.flow_date.strftime("%Y-%m-%d") if cf.flow_date else "",
            cf.amount, cf.currency or "EUR", (cf.description or "")[:500]
        ])

    # --- Hoja Ingresos ---
    ws_inc = wb.create_sheet("Ingresos", 1)
    inc_cats = IncomeCategory.query.filter_by(user_id=user_id).order_by(IncomeCategory.name).all()
    ws_inc.append(["--- Categorías ---"])
    ws_inc.append(["id_ref", "nombre", "icono", "parent_ref"])
    for c in inc_cats:
        parent_ref = c.parent.name if c.parent else ""
        ws_inc.append([f"inc_cat_{c.id}", c.name, c.icon or "💵", parent_ref])
    ws_inc.append([])
    ws_inc.append(["--- Ingresos ---"])
    ws_inc.append(["categoria", "importe", "descripcion", "fecha", "notas", "recurrente", "frecuencia"])
    for inc in Income.query.filter_by(user_id=user_id).order_by(Income.date).all():
        cat = inc.category
        cat_name = cat.name if cat else ""
        ws_inc.append([
            cat_name, inc.amount, (inc.description or "")[:500],
            inc.date.strftime("%Y-%m-%d") if inc.date else "",
            (inc.notes or "")[:500], inc.is_recurring, inc.recurrence_frequency or ""
        ])

    # --- Hoja Gastos ---
    ws_exp = wb.create_sheet("Gastos", 2)
    exp_cats = ExpenseCategory.query.filter_by(user_id=user_id).order_by(ExpenseCategory.name).all()
    ws_exp.append(["--- Categorías ---"])
    ws_exp.append(["id_ref", "nombre", "icono", "parent_ref"])
    for c in exp_cats:
        parent_ref = c.parent.name if c.parent else ""
        ws_exp.append([f"exp_cat_{c.id}", c.name, c.icon or "💰", parent_ref])
    ws_exp.append([])
    ws_exp.append(["--- Gastos ---"])
    ws_exp.append(["categoria", "importe", "descripcion", "fecha", "notas", "deuda_plan", "deuda_total", "deuda_meses", "deuda_inicio"])
    for exp in Expense.query.filter_by(user_id=user_id).order_by(Expense.date).all():
        cat = exp.category
        cat_name = cat.name if cat else ""
        dp = exp.debt_plan
        dp_name = dp.name if dp else ""
        dp_total = dp.total_amount if dp else ""
        dp_months = dp.months if dp else ""
        dp_start = dp.start_date.strftime("%Y-%m-%d") if dp and dp.start_date else ""
        ws_exp.append([
            cat_name, exp.amount, (exp.description or "")[:500],
            exp.date.strftime("%Y-%m-%d") if exp.date else "",
            (exp.notes or "")[:500], dp_name, dp_total, dp_months, dp_start
        ])

    # --- Hoja Deudas ---
    ws_debt = wb.create_sheet("Deudas", 3)
    ws_debt.append(["nombre", "total", "meses", "fecha_inicio", "estado", "categoria", "inmueble_direccion", "notas"])
    for dp in DebtPlan.query.filter_by(user_id=user_id).all():
        cat = dp.category
        pid = getattr(dp, 'property_id', None)
        prop = RealEstateProperty.query.get(pid) if pid else None
        ws_debt.append([
            dp.name, dp.total_amount, dp.months,
            dp.start_date.strftime("%Y-%m-%d") if dp.start_date else "",
            dp.status or "ACTIVE", cat.name if cat else "",
            prop.address if prop else "", (dp.notes or "")[:500]
        ])

    # --- Hoja Bancos ---
    ws_bank = wb.create_sheet("Bancos", 4)
    ws_bank.append(["nombre", "icono", "año", "mes", "saldo"])
    for b in Bank.query.filter_by(user_id=user_id).all():
        for bal in b.balances.order_by(BankBalance.year, BankBalance.month).all():
            ws_bank.append([b.name, b.icon or "🏦", bal.year, bal.month, bal.amount])

    # --- Hoja Inmuebles ---
    ws_re = wb.create_sheet("Inmuebles", 5)
    ws_re.append(["direccion", "tipo", "precio_compra", "fecha_compra", "notas"])
    for p in RealEstateProperty.query.filter_by(user_id=user_id).all():
        ws_re.append([
            p.address, p.property_type, p.purchase_price,
            p.purchase_date.strftime("%Y-%m-%d") if p.purchase_date else "",
            (p.notes or "")[:500]
        ])
    ws_re.append([])
    ws_re.append(["--- Tasaciones ---"])  # No usar = (Excel lo interpreta como fórmula)
    ws_re.append(["direccion", "año", "valor"])
    for p in RealEstateProperty.query.filter_by(user_id=user_id).all():
        for v in p.valuations.order_by(PropertyValuation.year).all():
            ws_re.append([p.address, v.year, v.value])

    # --- Hoja Watchlist (listas de seguimiento) ---
    ws_wl = wb.create_sheet("Watchlist", 6)
    ws_wl.append(["symbol", "isin", "next_earnings_date", "per_ntm", "ntm_dividend_yield", "eps", "cagr_revenue_yoy"])
    for w in Watchlist.query.filter_by(user_id=user_id).all():
        a = w.asset
        if a:
            ws_wl.append([
                a.symbol or "",
                a.isin or "",
                w.next_earnings_date.strftime("%Y-%m-%d") if w.next_earnings_date else "",
                w.per_ntm,
                w.ntm_dividend_yield,
                w.eps,
                w.cagr_revenue_yoy,
            ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def import_user_data(user_id: int, file_content: bytes) -> Tuple[bool, str, Dict[str, int]]:
    """
    Importa datos desde un .xlsx. Retorna (éxito, mensaje, estadísticas).
    No elimina datos existentes; añade/actualiza según corresponda.
    """
    from openpyxl import load_workbook

    stats = {"transacciones": 0, "cashflows": 0, "ingresos": 0, "gastos": 0, "deudas": 0, "bancos": 0, "inmuebles": 0, "tasaciones": 0, "watchlist": 0}
    cat_inc_by_name = {}
    cat_exp_by_name = {}
    props_by_addr = {}

    try:
        wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    except Exception as e:
        return False, f"Error al leer el archivo: {e}", stats

    broker_cache = {}
    account_cache = {}
    try:
        # --- Cargar o crear categorías de ingresos ---
        if "Ingresos" in wb.sheetnames:
            ws = wb["Ingresos"]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            sep = next((i for i, r in enumerate(rows) if r and "categorías" in str(r[0] or "").lower()), -1)
            if sep >= 0 and sep + 2 < len(rows):
                header = rows[sep + 1]
                for row in rows[sep + 2:]:
                    if not row or not row[1]:
                        continue
                    if row[0] and str(row[0]).strip().startswith("---"):
                        break
                    name = str(row[1]).strip()
                    if not name:
                        break
                    icon = str(row[2]).strip() if len(row) > 2 and row[2] else "💵"
                    parent_name = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                    parent_id = cat_inc_by_name.get(parent_name) if parent_name else None
                    cat = IncomeCategory.query.filter_by(user_id=user_id, name=name).first()
                    if not cat:
                        cat = IncomeCategory(user_id=user_id, name=name, icon=icon, parent_id=parent_id)
                        db.session.add(cat)
                        db.session.flush()
                    cat_inc_by_name[name] = cat.id

        # --- Cargar o crear categorías de gastos ---
        if "Gastos" in wb.sheetnames:
            ws = wb["Gastos"]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            sep = next((i for i, r in enumerate(rows) if r and "categorías" in str(r[0] or "").lower()), -1)
            if sep >= 0 and sep + 2 < len(rows):
                for row in rows[sep + 2:]:
                    if not row or not row[1]:
                        continue
                    if row[0] and str(row[0]).strip().startswith("---"):
                        break
                    name = str(row[1]).strip()
                    if not name:
                        break
                    icon = str(row[2]).strip() if len(row) > 2 and row[2] else "💰"
                    parent_name = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                    parent_id = cat_exp_by_name.get(parent_name) if parent_name else None
                    cat = ExpenseCategory.query.filter_by(user_id=user_id, name=name).first()
                    if not cat:
                        cat = ExpenseCategory(user_id=user_id, name=name, icon=icon, parent_id=parent_id)
                        db.session.add(cat)
                        db.session.flush()
                    cat_exp_by_name[name] = cat.id

        db.session.commit()

        # --- Importar transacciones ---
        if "Transacciones" in wb.sheetnames:
            ws = wb["Transacciones"]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            if rows and rows[0][0] == "broker":
                asset_cache = {}
                for row in rows[1:]:
                    if not row or not row[0]:
                        if row and str(row[0] or "").startswith("---"):
                            break
                        continue
                    if str(row[0]).startswith("---"):
                        break
                    broker_name = str(row[0]).strip() or "Manual"
                    account_name = str(row[1]).strip() or "Cuenta"
                    key_b = broker_name
                    key_a = f"{broker_name}|{account_name}"
                    if key_b not in broker_cache:
                        b = Broker.query.filter_by(name=broker_name).first()
                        if not b:
                            b = Broker(name=broker_name, full_name=broker_name)
                            db.session.add(b)
                            db.session.flush()
                        broker_cache[key_b] = b.id
                    if key_a not in account_cache:
                        acct = BrokerAccount.query.filter_by(
                            user_id=user_id,
                            broker_id=broker_cache[key_b]
                        ).filter(BrokerAccount.account_name == account_name).first()
                        if not acct:
                            acct = BrokerAccount(
                                user_id=user_id,
                                broker_id=broker_cache[key_b],
                                account_name=account_name,
                            )
                            db.session.add(acct)
                            db.session.flush()
                        account_cache[key_a] = acct.id
                    asset_type = str(row[2]).strip() if len(row) > 2 else ""
                    symbol = str(row[3]).strip() if len(row) > 3 else ""
                    isin = str(row[4]).strip() if len(row) > 4 else ""
                    if not asset_type and not symbol:
                        continue
                    asset = None
                    if symbol or isin:
                        if isin:
                            asset = Asset.query.filter_by(isin=isin).first()
                        if not asset and symbol:
                            asset = Asset.query.filter(Asset.symbol == symbol).first()
                        if not asset:
                            asset = Asset(
                                symbol=symbol or None,
                                isin=isin or None,
                                name=symbol or isin or "Activo",
                                asset_type=asset_type or "Stock",
                                currency="EUR",
                            )
                            db.session.add(asset)
                            db.session.flush()
                    if asset:
                        txn_date = _parse_date(row[6])
                        if not txn_date:
                            continue
                        from datetime import datetime as dt
                        txn = Transaction(
                            user_id=user_id,
                            account_id=account_cache[key_a],
                            asset_id=asset.id,
                            transaction_type=str(row[5]) or "BUY",
                            transaction_date=dt.combine(txn_date, dt.min.time()),
                            quantity=float(row[7]) if row[7] is not None else None,
                            price=float(row[8]) if row[8] is not None else None,
                            amount=float(row[9]) if row[9] is not None else 0,
                            currency=str(row[10]).strip() if row[10] else "EUR",
                            commission=float(row[11]) if row[11] is not None else 0,
                            fees=float(row[12]) if row[12] is not None else 0,
                            tax=float(row[13]) if row[13] is not None else 0,
                            description=(str(row[14]) or "")[:500],
                            notes=(str(row[15]) or "")[:500],
                            source=str(row[16]).strip() if len(row) > 16 and row[16] else "IMPORT",
                        )
                        db.session.add(txn)
                        stats["transacciones"] += 1

            # CashFlows
            cf_start = next((i for i, r in enumerate(rows) if r and "CashFlow" in str(r[0] or "")), -1)
            if cf_start >= 0 and cf_start + 2 < len(rows):
                for row in rows[cf_start + 2:]:
                    if not row or not row[0]:
                        break
                    broker_name = str(row[0]).strip() or "Manual"
                    account_name = str(row[1]).strip() or "Cuenta"
                    key_b = broker_name
                    key_a = f"{broker_name}|{account_name}"
                    if key_b not in broker_cache:
                        b = Broker.query.filter_by(name=broker_name).first()
                        if not b:
                            b = Broker(name=broker_name, full_name=broker_name)
                            db.session.add(b)
                            db.session.flush()
                        broker_cache[key_b] = b.id
                    if key_a not in account_cache:
                        acct = BrokerAccount.query.filter_by(user_id=user_id, broker_id=broker_cache[key_b]).filter(
                            BrokerAccount.account_name == account_name
                        ).first()
                        if not acct:
                            acct = BrokerAccount(user_id=user_id, broker_id=broker_cache[key_b], account_name=account_name)
                            db.session.add(acct)
                            db.session.flush()
                        account_cache[key_a] = acct.id
                    flow_date = _parse_date(row[3])
                    if flow_date:
                        cf = CashFlow(
                            user_id=user_id,
                            account_id=account_cache[key_a],
                            flow_type=str(row[2]) or "DEPOSIT",
                            flow_date=flow_date,
                            amount=float(row[4]) if row[4] is not None else 0,
                            currency=str(row[5]).strip() if row[5] else "EUR",
                            description=(str(row[6]) or "")[:500],
                        )
                        db.session.add(cf)
                        stats["cashflows"] += 1

        db.session.commit()

        # --- Importar Inmuebles y Tasaciones ---
        if "Inmuebles" in wb.sheetnames:
            ws = wb["Inmuebles"]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            prop_section = next((i for i, r in enumerate(rows) if r and "direccion" in str(r[0] or "").lower() and "tasaciones" not in str(r[0] or "")), -1)
            if prop_section >= 0:
                for row in rows[prop_section + 1:]:
                    if not row or not row[0]:
                        continue
                    addr = str(row[0]).strip()
                    if not addr or "---" in addr:
                        break
                    # Saltar fila de cabecera propiedades (precio_compra)
                    if str(row[2] or "").strip().lower() == "precio_compra":
                        continue
                    # Saltar fila de cabecera tasaciones (direccion, año, valor) tratada como inmueble
                    if addr.lower() == "direccion" and str(row[1] or "").strip().lower() == "año":
                        continue
                    p = RealEstateProperty.query.filter_by(user_id=user_id, address=addr).first()
                    if not p:
                        p = RealEstateProperty(
                            user_id=user_id,
                            address=addr,
                            property_type=str(row[1]).strip() if len(row) > 1 and row[1] else "casa",
                            purchase_price=_safe_float(row[2]) if len(row) > 2 else 0,
                            purchase_date=_parse_date(row[3]) or date.today(),
                            notes=(str(row[4]) or "")[:2000] if len(row) > 4 else None,
                        )
                        db.session.add(p)
                        db.session.flush()
                    props_by_addr[addr] = p
                    stats["inmuebles"] = len(props_by_addr)

            val_section = next((i for i, r in enumerate(rows) if r and "tasaciones" in str(r[0] or "").lower()), -1)
            if val_section >= 0 and val_section + 2 < len(rows):
                for row in rows[val_section + 2:]:
                    if not row or not row[0]:
                        break
                    addr = str(row[0]).strip()
                    if not addr:
                        continue
                    # Saltar fila de cabecera (direccion, año, valor)
                    if str(row[2] or "").strip().lower() == "valor" or str(row[1] or "").strip().lower() == "año":
                        continue
                    prop = props_by_addr.get(addr) or RealEstateProperty.query.filter_by(user_id=user_id, address=addr).first()
                    if prop and len(row) > 2:
                        year = _safe_float(row[1], -1)
                        val = _safe_float(row[2], -1)
                        if year < 1900 or year > 2100 or val < 0:
                            continue
                        year = int(year)
                        existing = PropertyValuation.query.filter_by(property_id=prop.id, year=year).first()
                        if not existing:
                            pv = PropertyValuation(property_id=prop.id, year=year, value=val)
                            db.session.add(pv)
                            stats["tasaciones"] += 1

        db.session.commit()

        # --- Importar Deudas (después de inmuebles para vincular) ---
        if "Deudas" in wb.sheetnames:
            ws = wb["Deudas"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for row in rows:
                if not row or not row[0]:
                    continue
                name = str(row[0]).strip()
                if name.lower() == "nombre":  # saltar cabecera
                    continue
                total = _safe_float(row[1])
                months = int(row[2]) if row[2] is not None else 12
                start = _parse_date(row[3]) or date.today()
                status = str(row[4]).strip() if len(row) > 4 and row[4] else "ACTIVE"
                cat_name = str(row[5]).strip() if len(row) > 5 and row[5] else "Otros"
                prop_addr = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                cat = ExpenseCategory.query.filter_by(user_id=user_id, name=cat_name).first()
                if not cat:
                    cat = ExpenseCategory(user_id=user_id, name=cat_name, icon="💰")
                    db.session.add(cat)
                    db.session.flush()
                prop_id = None
                if prop_addr:
                    prop = RealEstateProperty.query.filter_by(user_id=user_id, address=prop_addr).first()
                    if prop:
                        prop_id = prop.id
                # Permitir varios DebtPlan con el mismo nombre (ej. Mesa 250€ pagado + Mesa 450€ activo)
                existing = DebtPlan.query.filter_by(
                    user_id=user_id, name=name, total_amount=total,
                    start_date=start, status=status
                ).first()
                if not existing:
                    dp = DebtPlan(
                        user_id=user_id,
                        category_id=cat.id,
                        name=name,
                        total_amount=total,
                        months=months,
                        start_date=start,
                        status=status,
                        property_id=prop_id,
                        notes=(str(row[7]) or "")[:2000] if len(row) > 7 else None,
                    )
                    db.session.add(dp)
                    stats["deudas"] += 1

        db.session.commit()

        # --- Importar Bancos ---
        if "Bancos" in wb.sheetnames:
            ws = wb["Bancos"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            bank_by_name = {}
            for row in rows:
                if not row or not row[0]:
                    continue
                name = str(row[0]).strip()
                bank = bank_by_name.get(name)
                if not bank:
                    bank = Bank.query.filter_by(user_id=user_id, name=name).first()
                    if not bank:
                        bank = Bank(user_id=user_id, name=name, icon=str(row[1]).strip() if len(row) > 1 and row[1] else "🏦")
                        db.session.add(bank)
                        db.session.flush()
                    bank_by_name[name] = bank
                if len(row) > 4 and row[2] is not None and row[3] is not None and row[4] is not None:
                    if str(row[4]).strip().lower() == "saldo":  # saltar cabecera
                        continue
                    year = int(_safe_float(row[2], 0))
                    month = int(_safe_float(row[3], 0))
                    amount = _safe_float(row[4])
                    existing = BankBalance.query.filter_by(user_id=user_id, bank_id=bank.id, year=year, month=month).first()
                    if not existing:
                        bb = BankBalance(user_id=user_id, bank_id=bank.id, year=year, month=month, amount=amount)
                        db.session.add(bb)
                        stats["bancos"] += 1

        db.session.commit()

        # --- Importar Ingresos ---
        if "Ingresos" in wb.sheetnames:
            ws = wb["Ingresos"]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            inc_start = next((i for i, r in enumerate(rows) if r and "ingresos" in str(r[0] or "").lower()), -1)
            if inc_start >= 0 and inc_start + 2 < len(rows):
                for row in rows[inc_start + 2:]:
                    if not row or len(row) < 4:
                        continue
                    cat_name = str(row[0]).strip() if row[0] else ""
                    if not cat_name:
                        continue
                    cat_id = cat_inc_by_name.get(cat_name)
                    if not cat_id:
                        cat = IncomeCategory.query.filter_by(user_id=user_id, name=cat_name).first()
                        if not cat:
                            cat = IncomeCategory(user_id=user_id, name=cat_name, icon="💵")
                            db.session.add(cat)
                            db.session.flush()
                        cat_id = cat.id
                        cat_inc_by_name[cat_name] = cat_id
                    inc_date = _parse_date(row[3])
                    if inc_date:
                        inc = Income(
                            user_id=user_id,
                            category_id=cat_id,
                            amount=float(row[1]) if row[1] is not None else 0,
                            description=(str(row[2]) or "Importado")[:255],
                            date=inc_date,
                            notes=(str(row[4]) or "")[:2000] if len(row) > 4 else None,
                            is_recurring=bool(row[5]) if len(row) > 5 else False,
                            recurrence_frequency=(str(row[6]) or None) if len(row) > 6 and row[6] else None,
                        )
                        db.session.add(inc)
                        stats["ingresos"] += 1

        db.session.commit()

        # --- Importar Gastos ---
        if "Gastos" in wb.sheetnames:
            ws = wb["Gastos"]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            exp_start = next((i for i, r in enumerate(rows) if r and ("gastos" in str(r[0] or "").lower() or (str(r[0] or "").strip() == "categoria" and len(r) > 1 and "importe" in str(r[1] or "").lower()))), -1)
            data_start = exp_start + 2 if exp_start >= 0 and "gastos" in str(rows[exp_start][0] or "").lower() else exp_start + 1
            if exp_start >= 0 and data_start < len(rows):
                for row in rows[data_start:]:
                    if not row or len(row) < 4:
                        continue
                    if str(row[0] or "").strip().lower() == "categoria" and "importe" in str(row[1] or "").lower():
                        continue
                    cat_name = str(row[0]).strip() if row[0] else "Otros"
                    cat_id = cat_exp_by_name.get(cat_name)
                    if not cat_id:
                        cat = ExpenseCategory.query.filter_by(user_id=user_id, name=cat_name).first()
                        if not cat:
                            cat = ExpenseCategory(user_id=user_id, name=cat_name, icon="💰")
                            db.session.add(cat)
                            db.session.flush()
                        cat_id = cat.id
                        cat_exp_by_name[cat_name] = cat_id
                    exp_date = _parse_date(row[3])
                    if exp_date:
                        debt_plan_id = None
                        if len(row) > 5 and row[5]:
                            debt_plan_name = str(row[5]).strip()
                            debt_plan_total = _safe_float(row[6], None) if len(row) > 6 and row[6] is not None else None
                            debt_plan_months = int(_safe_float(row[7], 0)) if len(row) > 7 and row[7] is not None else None
                            debt_plan_start = _parse_date(row[8]) if len(row) > 8 and row[8] else None
                            q = DebtPlan.query.filter_by(user_id=user_id, name=debt_plan_name)
                            if debt_plan_total is not None and debt_plan_months and debt_plan_start:
                                q = q.filter_by(total_amount=debt_plan_total, months=debt_plan_months, start_date=debt_plan_start)
                            dp = q.first()
                            if dp:
                                debt_plan_id = dp.id
                        exp = Expense(
                            user_id=user_id,
                            category_id=cat_id,
                            amount=float(row[1]) if row[1] is not None else 0,
                            description=(str(row[2]) or "Importado")[:255],
                            date=exp_date,
                            notes=(str(row[4]) or "")[:2000] if len(row) > 4 else None,
                            debt_plan_id=debt_plan_id,
                        )
                        db.session.add(exp)
                        stats["gastos"] += 1

        db.session.commit()

        # --- Importar Watchlist ---
        if "Watchlist" in wb.sheetnames:
            ws = wb["Watchlist"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for row in rows:
                if not row or (not row[0] and not row[1]):
                    continue
                symbol = str(row[0]).strip() if len(row) > 0 and row[0] else ""
                isin = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                asset = None
                if isin:
                    asset = Asset.query.filter_by(isin=isin).first()
                if not asset and symbol:
                    asset = Asset.query.filter(Asset.symbol == symbol).first()
                if asset:
                    existing = Watchlist.query.filter_by(user_id=user_id, asset_id=asset.id).first()
                    if not existing:
                        wl = Watchlist(
                            user_id=user_id,
                            asset_id=asset.id,
                            next_earnings_date=_parse_date(row[2]) if len(row) > 2 else None,
                            per_ntm=_safe_float(row[3], None) if len(row) > 3 and row[3] is not None else None,
                            ntm_dividend_yield=_safe_float(row[4], None) if len(row) > 4 and row[4] is not None else None,
                            eps=_safe_float(row[5], None) if len(row) > 5 and row[5] is not None else None,
                            cagr_revenue_yoy=_safe_float(row[6], None) if len(row) > 6 and row[6] is not None else None,
                        )
                        db.session.add(wl)
                        stats["watchlist"] += 1

        db.session.commit()

        # Recalcular holdings tras importar transacciones
        if stats["transacciones"] or stats["cashflows"]:
            from app.services.portfolio_holding_service import recalculate_holdings
            for acct_id in BrokerAccount.query.filter_by(user_id=user_id).with_entities(BrokerAccount.id).all():
                recalculate_holdings(user_id, acct_id[0])
            from app.services.metrics.cache import MetricsCacheService
            MetricsCacheService.invalidate(user_id)

        db.session.commit()
        wb.close()
        total = sum(stats.values())
        msg = f"Importación completada. Elementos añadidos: {total} (Transacciones: {stats['transacciones']}, CashFlows: {stats['cashflows']}, Ingresos: {stats['ingresos']}, Gastos: {stats['gastos']}, Deudas: {stats['deudas']}, Bancos: {stats['bancos']}, Inmuebles: {stats['inmuebles']}, Tasaciones: {stats['tasaciones']}, Watchlist: {stats['watchlist']})"
        return True, msg, stats

    except Exception as e:
        db.session.rollback()
        try:
            wb.close()
        except Exception:
            pass
        return False, f"Error durante la importación: {str(e)}", stats


def _safe_float(val, default: Optional[float] = 0.0):
    """Convierte a float de forma segura; devuelve default si falla."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip().replace(",", "."))
    except (ValueError, TypeError):
        return default


def _parse_date(val) -> Optional[date]:
    """Convierte valor a date. Soporta datetime, str, y número serial de Excel."""
    if val is None:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    # Excel guarda fechas como número serial (días desde 1899-12-30)
    if isinstance(val, (int, float)):
        try:
            from datetime import datetime as dt, timedelta
            epoch = dt(1899, 12, 30).date()
            return epoch + timedelta(days=int(float(val)))
        except (ValueError, OverflowError):
            return None
    s = str(val).strip()
    if not s:
        return None
    try:
        from datetime import datetime as dt
        return dt.strptime(s[:10], "%Y-%m-%d").date()
    except Exception:
        try:
            return datetime.strptime(s[:10], "%d/%m/%Y").date()
        except Exception:
            return None
